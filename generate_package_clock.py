"""
FA Information Packet Generator — Clock-Hour Programs

Generates a multi-page financial aid packaging overview PDF for a single
student in a federal clock-hour program (HVAC, MSMA, etc.), then merges
selected institutional flyers into the same packet.

Federal Compliance Engine:
- Pell Grant distributed across two equal COD payment periods
- Direct Loan net disbursement computed after statutory origination fee
- Clock-hour regulatory callout per 34 CFR § 668.4(b)
- Interest capitalization modeling for unsubsidized/PLUS loans
- RAP income-driven repayment range modeling

Input: Excel workbook (see config/institutional.json for path)
Output: Combined PDF packet in named output folder
"""

import os
import json
import shutil
import math
import openpyxl
from pypdf import PdfWriter as PdfMerger
from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table,
    TableStyle, PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# ---------------------------------------------------------------------------
# CONFIG + PATH RESOLUTION
# ---------------------------------------------------------------------------

CONFIG_PATH = os.environ.get("INSTITUTION_CONFIG", "./config/institutional.json")
INPUT_EXCEL = os.environ.get("INPUT_EXCEL")
OUTPUT_ROOT = os.environ.get("OUTPUT_ROOT")
FLYERS_DIR = os.environ.get("FLYERS_DIR")

if not all([INPUT_EXCEL, OUTPUT_ROOT, FLYERS_DIR]):
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    inst = cfg["institution"]
    paths = cfg["paths"]
    flyer_map_rel = cfg["flyer_map"]
    program_meta = cfg["program_metadata"]
    INPUT_EXCEL = INPUT_EXCEL or paths["input_excel"]
    OUTPUT_ROOT = OUTPUT_ROOT or paths["output_root"]
    FLYERS_DIR = FLYERS_DIR or paths["flyers_directory"]
else:
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    inst = cfg["institution"]
    flyer_map_rel = cfg["flyer_map"]
    program_meta = cfg["program_metadata"]

FLYER_MAP = {k: os.path.join(FLYERS_DIR, v) for k, v in flyer_map_rel.items()}
os.makedirs(OUTPUT_ROOT, exist_ok=True)

NAVY = colors.HexColor(inst["primary_color_hex"])
GOLD = colors.HexColor(inst["accent_color_hex"])
DARK_GRAY = colors.HexColor("#333333")
LIGHT_BG = colors.HexColor("#F4F6F9")

# ---------------------------------------------------------------------------
# DATA LOAD
# ---------------------------------------------------------------------------

selected_flyers = []
student_name = "Unknown_Student"
program_name = "Unknown_Program"
program_cost = 0.0
pell_award = 0.0
sub_gross = 0.0
unsub_gross = 0.0
parent_plus_gross = 0.0
interest_rate_input = program_meta["default_student_interest_rate"]
plus_interest_rate_input = round(
    program_meta["default_student_interest_rate"]
    + program_meta["plus_interest_rate_spread"],
    4,
)
loan_total = 0.0
va_type = ""
q_costs = []


def safe_float(val):
    if val is None:
        return 0.0
    s = str(val).strip()
    if s.startswith("#") or not s:
        return 0.0
    try:
        for char in ["$", ",", " ", "%"]:
            s = s.replace(char, "")
        return float(s)
    except ValueError:
        return 0.0


def _read_rate(cell_value, fallback):
    raw = safe_float(cell_value)
    if raw > 1.0:
        return raw / 100.0
    if raw > 0:
        return raw
    return fallback


temp_snapshot = "_temp_snapshot.xlsx"
try:
    shutil.copy2(INPUT_EXCEL, temp_snapshot)
    wb = openpyxl.load_workbook(temp_snapshot, data_only=True, read_only=True)

    if "Package_Data" in wb.sheetnames:
        ws = wb["Package_Data"]
        student_name = str(ws["B1"].value).strip() if ws["B1"].value else "Unknown_Student"
        program_name = str(ws["B2"].value).strip() if ws["B2"].value else "Unknown_Program"
        program_cost = safe_float(ws["B3"].value)
        pell_award = safe_float(ws["B4"].value)
        sub_gross = safe_float(ws["B5"].value)
        unsub_gross = safe_float(ws["B6"].value)
        interest_rate_input = _read_rate(ws["B7"].value, program_meta["default_student_interest_rate"])
        parent_plus_gross = safe_float(ws["B8"].value)
        plus_interest_rate_input = _read_rate(
            ws["B9"].value,
            round(program_meta["default_student_interest_rate"] + program_meta["plus_interest_rate_spread"], 4),
        )
        loan_total = safe_float(ws["B10"].value) or (sub_gross + unsub_gross + parent_plus_gross)

        if ws["F4"].value is not None:
            va_type = str(ws["F4"].value).strip()

        for row in range(2, 8):
            flyer_name = ws.cell(row=row, column=4).value
            checked = ws.cell(row=row, column=5).value
            if flyer_name is not None and checked is not None:
                if str(checked).strip().upper() in ["X", "YES", "TRUE", "1", "1.0", "CHECK"]:
                    selected_flyers.append(str(flyer_name).strip())

    if "Program_Master" in wb.sheetnames:
        pm_ws = wb["Program_Master"]
        for row in range(2, 25):
            row_prog = str(pm_ws.cell(row=row, column=1).value).strip().upper()
            if row_prog == program_name.upper().strip():
                raw = [safe_float(pm_ws.cell(row=row, column=c).value) for c in (3, 4, 5, 6)]
                q_costs = [c for c in raw if c > 0]
                break

    wb.close()
finally:
    if os.path.exists(temp_snapshot):
        os.remove(temp_snapshot)

if student_name == "nan" or not student_name or student_name.startswith("#"):
    student_name = "Unknown_Student"

is_va_student = "VA Student" in selected_flyers
va_label = (
    f"Chapter {'33' if '33' in va_type else '35'} "
    f"{'Post-9/11' if '33' in va_type else 'DEA'}"
    if is_va_student else ""
)

# ---------------------------------------------------------------------------
# COD ENGINE MATH (INTEGER-BASED PER-DISBURSEMENT ALLOCATION)
# ---------------------------------------------------------------------------

origination_fee_rate = program_meta["origination_fee_student_loans"]
plus_fee_rate = program_meta["origination_fee_fee_parent_plus"] if "origination_fee_parent_plus" in program_meta else program_meta["origination_fee_parent_plus"]
plus_fee_rate = program_meta["origination_fee_parent_plus"]

if not q_costs:
    q_costs = [program_cost / 2.0, program_cost / 2.0]

total_terms = len(q_costs)

# Pell: split evenly, ceiling then floor to land on exact total
pell_p1 = math.ceil(pell_award / 2.0)
pell_p2 = math.floor(pell_award / 2.0)

# Per-disbursement net loan values after origination fee deduction
sub_disb_net = math.ceil((sub_gross / 2.0) * (1.0 - origination_fee_rate)) if sub_gross > 0 else 0.0
unsub_disb_net = math.ceil((unsub_gross / 2.0) * (1.0 - origination_fee_rate)) if unsub_gross > 0 else 0.0
plus_disb_net = math.ceil((parent_plus_gross / 2.0) * (1.0 - plus_fee_rate)) if parent_plus_gross > 0 else 0.0

total_loan_disb_net = sub_disb_net + unsub_disb_net + plus_disb_net
student_loan_gross = sub_gross + unsub_gross
total_net_applied_funding = (pell_p1 + pell_p2) + (total_loan_disb_net * 2)

if is_va_student and "33" in va_type:
    loan_total = 0.0
    student_loan_gross = 0.0
    parent_plus_gross = 0.0
    sub_disb_net = 0.0
    unsub_disb_net = 0.0
    plus_disb_net = 0.0
    total_loan_disb_net = 0.0
    total_net_applied_funding = pell_award

# Distribute disbursements across term milestones
pell_posts = [0.0] * total_terms
loan_posts = [0.0] * total_terms

if total_terms >= 1:
    pell_posts[0] = pell_p1
    loan_posts[0] = total_loan_disb_net
if total_terms >= 3:
    pell_posts[2] = pell_p2
    loan_posts[2] = total_loan_disb_net
elif total_terms == 2:
    pell_posts[1] = pell_p2
    loan_posts[1] = total_loan_disb_net

timeline_ledger = []
running_balance = 0.0

for i in range(total_terms):
    term_charge = round(q_costs[i], 2)
    term_pell = round(pell_posts[i], 2)
    term_loan = round(loan_posts[i], 2)

    if is_va_student and "33" in va_type:
        running_balance = round(running_balance + term_charge - (term_charge + term_pell), 2)
    elif is_va_student and "35" in va_type:
        running_balance = round(running_balance + term_charge - term_pell, 2)
    else:
        running_balance = round(running_balance + term_charge - (term_pell + term_loan), 2)

    if abs(running_balance) <= 1.0:
        display_balance = 0.0
    elif running_balance < 0:
        display_balance = -float(math.ceil(abs(running_balance)))
    else:
        display_balance = float(math.ceil(running_balance))

    timeline_ledger.append({
        "label": f"Term / Quarter {i+1} Start",
        "charge": term_charge,
        "pell": term_pell,
        "loan": term_loan,
        "balance": display_balance,
    })

balance_refund = timeline_ledger[-1]["balance"]
ch33_refund_value = pell_award

# ---------------------------------------------------------------------------
# COVER SHEET LAYOUT (PAGE 1)
# ---------------------------------------------------------------------------

cover_pdf = "_temp_cover.pdf"
doc = SimpleDocTemplate(
    cover_pdf, pagesize=letter,
    leftMargin=36, rightMargin=36, topMargin=30, bottomMargin=30,
)
styles = getSampleStyleSheet()
story = []

title_style = ParagraphStyle(
    "TStyle", fontName="Helvetica-Bold", fontSize=20, leading=24,
    textColor=colors.white, alignment=1, spaceAfter=4,
)
subtitle_style = ParagraphStyle(
    "SubStyle", fontName="Helvetica-Bold", fontSize=13, leading=16,
    textColor=GOLD, alignment=1, spaceAfter=2,
)
program_style = ParagraphStyle(
    "PStyle", fontName="Helvetica", fontSize=10.5, leading=14,
    textColor=colors.white, alignment=1,
)
section_style = ParagraphStyle(
    "SecStyle", fontName="Helvetica-Bold", fontSize=13, leading=16,
    textColor=NAVY, spaceBefore=12, spaceAfter=4,
)
th_style = ParagraphStyle(
    "THStyle", fontName="Helvetica-Bold", fontSize=9.5, leading=12,
    textColor=colors.white, alignment=1,
)
td_style = ParagraphStyle(
    "TDStyle", fontName="Helvetica", fontSize=9.5, leading=12,
    textColor=DARK_GRAY, alignment=1,
)
td_left = ParagraphStyle(
    "TDLeft", fontName="Helvetica", fontSize=9.5, leading=12,
    textColor=DARK_GRAY, alignment=0,
)
outcome_style = ParagraphStyle(
    "OutcomeStyle", fontName="Helvetica", fontSize=10.5,
    textColor=DARK_GRAY, leading=14,
)

funding_label = (
    f"Total Federal Grants: ${total_net_applied_funding:,.2f} ({va_label})"
    if is_va_student
    else f"Total Estimated Net Financial Aid: ${total_net_applied_funding:,.2f}"
)

header_data = [
    [Paragraph("FINANCIAL AID PACKAGING OVERVIEW", title_style)],
    [Paragraph(
        f"Prepared For: {student_name} {f'({va_label})' if is_va_student else ''}",
        subtitle_style,
    )],
    [Paragraph(
        f"Program: {program_name} | Program Cost: ${program_cost:,.2f} | {funding_label}",
        program_style,
    )],
]
header_table = Table(header_data, colWidths=[540])
header_table.setStyle(TableStyle([
    ("BACKGROUND", (0, 0), (-1, -1), NAVY),
    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ("TOPPADDING", (0, 0), (-1, -1), 12),
    ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
]))
story.append(header_table)
story.append(Spacer(1, 8))

# Outcome block
if is_va_student and "33" in va_type:
    outcome_text = (
        f"<b>CHAPTER 33 POST-9/11 SUMMARY:</b> Tuition and mandatory fees are certified and paid "
        f"directly to {inst['short_name']} by the VA. Your Federal Pell Grant is applied to your "
        f"ledger, creating an estimated credit balance refund of <b>${ch33_refund_value:,.2f}</b>. "
        f"Any Title IV credit balance is processed and refunded to the student within "
        f"{program_meta['credit_balance_refund_days']} calendar days of occurring on your student "
        f"account ledger (34 CFR § 668.164(h))."
    )
    box_bg = colors.HexColor("#E2F0D9")
    box_line = colors.HexColor("#385723")
elif is_va_student and "35" in va_type:
    ch35_real_balance = program_cost - pell_award
    outcome_text = (
        f"<b>CHAPTER 35 DEA SUMMARY:</b> The VA distributes monthly allowance stipends "
        f"<b>directly to the student</b>, not to the school. Your Federal Pell Grant reduces "
        f"direct program costs on your institutional ledger, leaving an estimated remaining "
        f"balance of <b>${ch35_real_balance:,.2f}</b>. An Institutional Payment Arrangement "
        f"is required to establish manageable monthly installments."
    )
    box_bg = colors.HexColor("#FCE4D6")
    box_line = colors.HexColor("#C65911")
else:
    if balance_refund <= 0:
        outcome_text = (
            f"<b>ESTIMATED CREDIT BALANCE:</b> Your net financial aid awards completely clear "
            f"your direct program costs, leaving an estimated credit balance refund of "
            f"<b>${abs(balance_refund):,.2f}</b>. In accordance with 34 CFR § 668.164(h), any "
            f"Title IV credit balance is processed and issued to the student within "
            f"{program_meta['credit_balance_refund_days']} calendar days of occurring on your "
            f"active student account ledger."
        )
        box_bg = colors.HexColor("#E2F0D9")
        box_line = colors.HexColor("#385723")
    else:
        outcome_text = (
            f"<b>ESTIMATED OUT-OF-POCKET BALANCE DUE:</b> After applying your net financial aid "
            f"disbursements against institutional program costs, an estimated balance of "
            f"<b>${balance_refund:,.2f}</b> remains. <br/><br/><b>Payment Options Available:</b> "
            f"A customized Institutional Payment Plan will be arranged with the Business Office "
            f"to establish monthly installment options."
        )
        box_bg = colors.HexColor("#FCE4D6")
        box_line = colors.HexColor("#C65911")

outcome_table = Table([[Paragraph(outcome_text, outcome_style)]], colWidths=[540])
outcome_table.setStyle(TableStyle([
    ("BACKGROUND", (0, 0), (-1, -1), box_bg),
    ("BOX", (0, 0), (-1, -1), 1.2, box_line),
    ("TOPPADDING", (0, 0), (-1, -1), 10),
    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ("LEFTPADDING", (0, 0), (-1, -1), 10),
    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
]))
story.append(outcome_table)

story.append(Paragraph("Account Ledger Timeline", section_style))

timeline_data = [[
    Paragraph("Term Milestone", th_style),
    Paragraph("Tuition & Fees", th_style),
    Paragraph("Pell Grant", th_style),
    Paragraph("Net Loan Funds (All)", th_style),
    Paragraph("Estimated Balance", th_style),
]]
for row in timeline_ledger:
    if is_va_student and "33" in va_type:
        bal_str = Paragraph("<font color='#385723'><b>VA Certified</b></font>", td_style)
    else:
        bal_str = Paragraph(
            f"-${abs(row['balance']):,.2f} (Estimated Credit)" if row["balance"] < 0
            else f"${row['balance']:,.2f}",
            td_style,
        )
    timeline_data.append([
        Paragraph(row["label"], td_left),
        Paragraph(f"${row['charge']:,.2f}", td_style),
        Paragraph(f"${row['pell']:,.2f}" if row["pell"] > 0 else "$0.00", td_style),
        Paragraph(f"${row['loan']:,.2f}" if row["loan"] > 0 else "$0.00", td_style),
        bal_str,
    ])

timeline_table = Table(timeline_data, colWidths=[130, 105, 100, 100, 105])
timeline_table.setStyle(TableStyle([
    ("BACKGROUND", (0, 0), (-1, 0), NAVY),
    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
    ("TOPPADDING", (0, 0), (-1, -1), 6),
    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_BG]),
]))
story.append(timeline_table)

# Clock-hour regulatory callout
prog_upper = program_name.upper()
prog_hours = "880" if "HVAC" in prog_upper else "891"
hour_milestone = (
    "440 clock hours and 22 weeks" if "HVAC" in prog_upper
    else "445.5 clock hours and 22 weeks"
)
gate_warning_text = (
    f"⚠️ <b>CLOCK-HOUR ENROLLMENT REGULATORY DISCLOSURE:</b> This program operates on a federal "
    f"clock-hour framework (Total Program Hours: {prog_hours}). While institutional costs are "
    f"assessed on a term structure, federal financial aid is divided into two equal payment "
    f"period disbursements. The first disbursement is scheduled at the start of your program. "
    f"The second disbursement <b>WILL NOT disburse until you have successfully completed "
    f"{hour_milestone} of instructional time</b> (34 CFR § 668.4(b)). Attendance directly "
    f"dictates disbursement timing."
)
gate_table = Table(
    [[Paragraph(gate_warning_text, ParagraphStyle(
        "GateStyle", fontName="Helvetica-Oblique", fontSize=8.5, leading=11,
        textColor=colors.HexColor("#7F6000"),
    ))]],
    colWidths=[540],
)
gate_table.setStyle(TableStyle([
    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF2CC")),
    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D6B656")),
    ("TOPPADDING", (0, 0), (-1, -1), 5),
    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
]))
story.append(Spacer(1, 4))
story.append(gate_table)
story.append(Spacer(1, 6))

# Loan obligation block
if is_va_student and "33" in va_type:
    loan_box_text = """
    <b>MILITARY EDUCATION BENEFITS POLICY: ESTIMATED $0.00 LOAN DEBT OFFERED</b><br/>
    <font size='9' color='#333333'>Based on current packaging, your institutional tuition and
    fees are covered fully via Chapter 33 veterans' benefits and non-repayable Federal Pell
    Grants; you currently have no federal student loan debt obligations scheduled for this cycle.</font>
    """
elif is_va_student and "35" in va_type:
    loan_box_text = f"""
    <b>CHAPTER 35 REPAYMENT &amp; LEDGER DISCLOSURE: PAYMENT ARRANGEMENT PLAN</b><br/>
    <font size='9' color='#333333'>The VA does not pay the school directly under Chapter 35.
    Federal Pell Grants apply directly to your institutional bill; however, an institutional
    payment arrangement plan is required to address any remaining balance.</font>
    """
else:
    plus_chunk = ""
    if parent_plus_gross > 0:
        plus_chunk = f"""<br/>
        &bull; <b>Parent PLUS Loan Scheduled:</b> Gross: <b>${parent_plus_gross:,.2f}</b> |
        Net per Disb: <b>${plus_disb_net:,.2f}</b> | Total Net Applied:
        <b>${plus_disb_net * 2:,.2f}</b><br/>
        <font color='#990000'><b>Important Note on Parent PLUS:</b> Parent PLUS loans are
        borrowed by and are the sole legal repayment obligation of the parent borrower,
        not the student.</font>
        """
    loan_box_text = f"""
    <b>TOTAL ESTIMATED GROSS LOAN FINANCING SCHEDULED: ${loan_total:,.2f}</b><br/>
    <font size='8.5' color='#333333'>Direct federal loans carry mandatory origination fees
    deducted by the U.S. Department of Education prior to ledger disbursement.</font><br/>
    <font size='8' color='#555555'>
    &bull; <b>Direct Student Loans (Sub/Unsub):</b> Gross Scheduled:
    <b>${student_loan_gross:,.2f}</b> | Net per Disb:
    <b>${sub_disb_net + unsub_disb_net:,.2f}</b> | Total Net Applied:
    <b>${(sub_disb_net + unsub_disb_net) * 2:,.2f}</b> (Student Repayment Obligation){plus_chunk}
    </font>
    """

loan_box_style = ParagraphStyle(
    "LoanBoxStyle", fontName="Helvetica", fontSize=10, textColor=NAVY, leading=13,
)
loan_box_table = Table([[Paragraph(loan_box_text, loan_box_style)]], colWidths=[540])
loan_box_table.setStyle(TableStyle([
    ("BACKGROUND", (0, 0), (-1, -1), LIGHT_BG),
    ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#002D62")),
    ("TOPPADDING", (0, 0), (-1, -1), 8),
    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ("LEFTPADDING", (0, 0), (-1, -1), 10),
    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
]))
story.append(loan_box_table)

story.append(Paragraph("Critical Disclosures", section_style))
disc_style = ParagraphStyle(
    "DiscStyle", fontName="Helvetica", fontSize=9, textColor=DARK_GRAY,
    leading=13, leftIndent=15, firstLineIndent=-10,
)
if is_va_student and "33" in va_type:
    d1 = "<b>VA Funding Timelines:</b> Post-9/11 tuition disbursements dispatch on a distinct federal timeline and post directly to the institution independently from Pell Grant schedules."
    d2 = "<b>Pell Grant Autonomy:</b> Your Pell Grant eligibility remains independent from GI Bill restrictions and applies directly to your institutional charges."
    d3 = f"<b>Credit Balance Processing:</b> Eligible Title IV credit balance refunds are processed and refunded within {program_meta['credit_balance_refund_days']} calendar days of posting to the ledger (34 CFR § 668.164(h))."
elif is_va_student and "35" in va_type:
    d1 = "<b>Direct Student Payments:</b> The VA will distribute Chapter 35 monthly allowance payments directly to your personal bank account, not to the school."
    d2 = "<b>Ledger Application:</b> Pell Grants apply to your institutional account first, reducing your immediate balance due to the business office."
    d3 = "<b>Payment Arrangements:</b> Please connect with the Financial Aid Office to establish an approved payment plan to resolve your remaining institutional balance."
else:
    d1 = f"<b>True Net Values:</b> Timeline schedule figures reflect net cash applied after statutory federal loan origination fees ({origination_fee_rate*100:.3f}% Direct Student / {plus_fee_rate*100:.3f}% Parent PLUS)."
    d2 = f"<b>30-Day Delay Rule:</b> First-time, first-year student borrowers are subject to a mandatory federal {program_meta['first_borrower_disbursement_delay_days']}-day delayed disbursement hold from the first day of classes (34 CFR § 685.303(b)(5))."
    d3 = f"<b>Credit Balance Distribution:</b> Any Title IV credit balance is refunded within {program_meta['credit_balance_refund_days']} calendar days of occurrence on the ledger in accordance with 34 CFR § 668.164(h)."

story.append(Paragraph(f"&bull; {d1}", disc_style))
story.append(Spacer(1, 3))
story.append(Paragraph(f"&bull; {d2}", disc_style))
story.append(Spacer(1, 3))
story.append(Paragraph(f"&bull; {d3}", disc_style))

# ---------------------------------------------------------------------------
# REPAYMENT MODELING (PAGE 2)
# ---------------------------------------------------------------------------

if (student_loan_gross + parent_plus_gross) > 0 and not (is_va_student and "33" in va_type):
    story.append(PageBreak())

    monthly_rate = interest_rate_input / 12.0
    num_payments_std = program_meta["standard_repayment_term_months"]
    in_school_grace_months = 16

    if sub_gross > 0:
        sub_repay_principal = sub_gross
        sub_std_monthly = sub_repay_principal * (
            (monthly_rate * ((1 + monthly_rate) ** num_payments_std))
            / (((1 + monthly_rate) ** num_payments_std) - 1)
        )
        sub_std_total_paid = sub_std_monthly * num_payments_std
    else:
        sub_repay_principal = sub_std_monthly = sub_std_total_paid = 0.0

    unsub_accrued_interest = unsub_gross * interest_rate_input * (in_school_grace_months / 12.0)
    unsub_repay_principal = unsub_gross + unsub_accrued_interest
    if unsub_repay_principal > 0:
        unsub_std_monthly = unsub_repay_principal * (
            (monthly_rate * ((1 + monthly_rate) ** num_payments_std))
            / (((1 + monthly_rate) ** num_payments_std) - 1)
        )
        unsub_std_total_paid = unsub_std_monthly * num_payments_std
    else:
        unsub_std_monthly = unsub_std_total_paid = 0.0

    student_combined_monthly = sub_std_monthly + unsub_std_monthly
    student_combined_total_paid = sub_std_total_paid + unsub_std_total_paid

    sub_rap_str = (
        f"<b>${program_meta['rap_minimum_payment']:.2f} – ${sub_std_monthly:,.2f}</b> / mo*"
        f"<br/><font size='6' color='#555555'>(${program_meta['rap_minimum_payment']:.2f}/mo minimum)</font>"
        if sub_repay_principal > 0 else "$0.00 / mo"
    )
    unsub_rap_str = (
        f"<b>${program_meta['rap_minimum_payment']:.2f} – ${unsub_std_monthly:,.2f}</b> / mo*"
        f"<br/><font size='6' color='#555555'>(${program_meta['rap_minimum_payment']:.2f}/mo minimum)</font>"
        if unsub_repay_principal > 0 else "$0.00 / mo"
    )

    if parent_plus_gross > 0:
        plus_monthly_rate = plus_interest_rate_input / 12.0
        plus_accrued_interest = parent_plus_gross * plus_interest_rate_input * (in_school_grace_months / 12.0)
        plus_repay_principal = parent_plus_gross + plus_accrued_interest
        plus_std_monthly = plus_repay_principal * (
            (plus_monthly_rate * ((1 + plus_monthly_rate) ** num_payments_std))
            / (((1 + plus_monthly_rate) ** num_payments_std) - 1)
        )
        plus_std_total_paid = plus_std_monthly * num_payments_std
    else:
        plus_accrued_interest = plus_repay_principal = plus_std_monthly = plus_std_total_paid = 0.0

    page2_title_style = ParagraphStyle(
        "P2Title", fontName="Helvetica-Bold", fontSize=15, leading=19,
        textColor=NAVY, spaceAfter=2,
    )
    page2_sub_style = ParagraphStyle(
        "P2Sub", fontName="Helvetica-Bold", fontSize=9.5, leading=13,
        textColor=GOLD, spaceAfter=6,
    )
    table_head_style = ParagraphStyle(
        "THHead", fontName="Helvetica-Bold", fontSize=8, leading=10,
        textColor=colors.white, alignment=1,
    )
    table_cell_style = ParagraphStyle(
        "TDCell", fontName="Helvetica", fontSize=8, leading=10,
        textColor=DARK_GRAY, alignment=1,
    )
    table_cell_left = ParagraphStyle(
        "TDLeft2", fontName="Helvetica", fontSize=8, leading=10,
        textColor=DARK_GRAY, alignment=0,
    )
    box_header_style = ParagraphStyle(
        "BHStyle", fontName="Helvetica-Bold", fontSize=9, leading=12,
        textColor=NAVY, spaceBefore=3, spaceAfter=2,
    )

    story.append(Paragraph("Federal Student &amp; Parent Loan Repayment Modeling", page2_title_style))
    story.append(Paragraph(
        f"Undergraduate Direct Loan Rate: <b>{interest_rate_input*100:.2f}%</b> | "
        f"Parent PLUS Rate: <b>{plus_interest_rate_input*100:.2f}%</b> | "
        f"Modeling Period: <b>{in_school_grace_months} Months</b>",
        page2_sub_style,
    ))

    # 1. Subsidized
    story.append(Paragraph(
        "1. Direct Subsidized Loan (Student Obligation &bull; Interest Covered by Government During School &amp; Grace)",
        box_header_style,
    ))
    sub_table_data = [
        [Paragraph("Repayment Plan", table_head_style),
         Paragraph("Original Principal", table_head_style),
         Paragraph("Accrued / Capitalized Interest", table_head_style),
         Paragraph("Repayment Balance", table_head_style),
         Paragraph("Est. Monthly Payment", table_head_style),
         Paragraph("Total Amount Paid", table_head_style)],
        [Paragraph("<b>Standard 10-Year</b><br/><font size='6.5' color='#555555'>120 Fixed Payments</font>", table_cell_left),
         Paragraph(f"${sub_gross:,.2f}", table_cell_style),
         Paragraph("$0.00<br/><font size='6.5' color='#274E13'><b>(100% Subsidized)</b></font>", table_cell_style),
         Paragraph(f"${sub_repay_principal:,.2f}", table_cell_style),
         Paragraph(f"<b>${sub_std_monthly:,.2f}</b> / mo", table_cell_style),
         Paragraph(f"${sub_std_total_paid:,.2f}", table_cell_style)],
        [Paragraph("<b>RAP (IDR Plan)</b><br/><font size='6.5' color='#555555'>Income-Driven Scale</font>", table_cell_left),
         Paragraph(f"${sub_gross:,.2f}", table_cell_style),
         Paragraph("$0.00<br/><font size='6.5' color='#274E13'><b>(100% Subsidized)</b></font>", table_cell_style),
         Paragraph(f"${sub_repay_principal:,.2f}", table_cell_style),
         Paragraph(sub_rap_str, table_cell_style),
         Paragraph("Income-Dependent<br/><font size='6.5' color='#555555'>(Up to 30 Yrs)</font>", table_cell_style)],
    ]
    sub_table = Table(sub_table_data, colWidths=[105, 80, 95, 80, 85, 95])
    sub_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_BG]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(sub_table)
    story.append(Spacer(1, 4))

    # 2. Unsubsidized
    story.append(Paragraph(
        "2. Direct Unsubsidized Loan (Student Obligation &bull; Interest Accrues from Disbursement)",
        box_header_style,
    ))
    unsub_table_data = [
        [Paragraph("Repayment Plan", table_head_style),
         Paragraph("Original Principal", table_head_style),
         Paragraph("Accrued In-School Interest", table_head_style),
         Paragraph("Capitalized Balance", table_head_style),
         Paragraph("Est. Monthly Payment", table_head_style),
         Paragraph("Total Amount Paid", table_head_style)],
        [Paragraph("<b>Standard 10-Year</b><br/><font size='6.5' color='#555555'>120 Fixed Payments</font>", table_cell_left),
         Paragraph(f"${unsub_gross:,.2f}", table_cell_style),
         Paragraph(f"${unsub_accrued_interest:,.2f}<br/><font size='6.5' color='#C65911'><b>(Capitalizes)</b></font>", table_cell_style),
         Paragraph(f"${unsub_repay_principal:,.2f}", table_cell_style),
         Paragraph(f"<b>${unsub_std_monthly:,.2f}</b> / mo", table_cell_style),
         Paragraph(f"${unsub_std_total_paid:,.2f}", table_cell_style)],
        [Paragraph("<b>RAP (IDR Plan)</b><br/><font size='6.5' color='#555555'>Income-Driven Scale</font>", table_cell_left),
         Paragraph(f"${unsub_gross:,.2f}", table_cell_style),
         Paragraph(f"${unsub_accrued_interest:,.2f}<br/><font size='6.5' color='#C65911'><b>(Capitalizes)</b></font>", table_cell_style),
         Paragraph(f"${unsub_repay_principal:,.2f}", table_cell_style),
         Paragraph(unsub_rap_str, table_cell_style),
         Paragraph("Income-Dependent<br/><font size='6.5' color='#555555'>(Up to 30 Yrs)</font>", table_cell_style)],
    ]
    unsub_table = Table(unsub_table_data, colWidths=[105, 80, 95, 80, 85, 95])
    unsub_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_BG]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(unsub_table)
    story.append(Spacer(1, 4))

    # 3. Parent PLUS (conditional)
    if parent_plus_gross > 0:
        story.append(Paragraph(
            f"3. Direct Parent PLUS Loan (<font color='#990000'><b>Parent Legal Repayment Obligation Only</b></font>)",
            box_header_style,
        ))
        plus_table_data = [
            [Paragraph("Repayment Plan", table_head_style),
             Paragraph("Original Principal", table_head_style),
             Paragraph("Accrued In-School Interest", table_head_style),
             Paragraph("Capitalized Balance", table_head_style),
             Paragraph("Est. Monthly Payment", table_head_style),
             Paragraph("Total Amount Paid", table_head_style)],
            [Paragraph("<b>Standard 10-Year</b><br/><font size='6.5' color='#555555'>120 Fixed Payments</font>", table_cell_left),
             Paragraph(f"${parent_plus_gross:,.2f}", table_cell_style),
             Paragraph(f"${plus_accrued_interest:,.2f}<br/><font size='6.5' color='#C65911'><b>(If Deferred)</b></font>", table_cell_style),
             Paragraph(f"${plus_repay_principal:,.2f}", table_cell_style),
             Paragraph(f"<b>${plus_std_monthly:,.2f}</b> / mo", table_cell_style),
             Paragraph(f"${plus_std_total_paid:,.2f}", table_cell_style)],
            [Paragraph("<b>RAP (IDR Plan)</b>", table_cell_left),
             Paragraph(f"${parent_plus_gross:,.2f}", table_cell_style),
             Paragraph("—", table_cell_style),
             Paragraph("—", table_cell_style),
             Paragraph("<font color='#990000'><b>Not Eligible</b></font><br/><font size='6' color='#555555'>Direct PLUS Ineligible</font>", table_cell_style),
             Paragraph("N/A<br/><font size='6' color='#555555'>(Standard/ICR Only)</font>", table_cell_style)],
        ]
        plus_table = Table(plus_table_data, colWidths=[105, 80, 95, 80, 85, 95])
        plus_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#660000")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_BG]),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(plus_table)
        story.append(Spacer(1, 4))

    if parent_plus_gross > 0:
        summary_text = (
            f"<b>FINANCING SUMMARY:</b> Student Loan Standard Payment: "
            f"<b>${student_combined_monthly:,.2f}/mo</b> | Parent PLUS Standard Payment: "
            f"<b>${plus_std_monthly:,.2f}/mo</b> (Total Combined Monthly: "
            f"${student_combined_monthly + plus_std_monthly:,.2f}/mo)"
        )
    else:
        summary_text = (
            f"<b>COMBINED REPAYMENT SUMMARY:</b> Total Scheduled Principal: "
            f"<b>${student_loan_gross:,.2f}</b> | Combined Standard Monthly Payment: "
            f"<b>${student_combined_monthly:,.2f} / month</b> (Total Lifecycle Cost: "
            f"${student_combined_total_paid:,.2f})"
        )

    summary_table = Table(
        [[Paragraph(summary_text, ParagraphStyle(
            "SumP", fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=NAVY,
        ))]],
        colWidths=[540],
    )
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF2CC")),
        ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#D6B656")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 4))

    plus_extra_disc = ""
    if parent_plus_gross > 0:
        plus_extra_disc = (
            "<br/>&bull; <b>Parent PLUS Loan Terms:</b> Parent PLUS loans are the sole legal "
            "responsibility of the parent borrower and cannot be transferred to the student. "
            "PLUS loans are not eligible for RAP. If an income-driven plan is required by the "
            "parent, the loan must first be consolidated into a Federal Direct Consolidation "
            "Loan to access the Income-Contingent Repayment (ICR) plan."
        )

    repay_disclosures = f"""
    <b>REPAYMENT PLAN DISCLOSURES &amp; REGULATORY NOTICE:</b><br/>
    <font size='7.5' color='#444444'>
    &bull; <b>Interest Capitalization:</b> Unsubsidized and Parent PLUS loans accrue interest
    daily from disbursement. Unpaid accrued interest is added (capitalized) to principal balances
    when entering active repayment.<br/>
    &bull; <b>Standard 10-Year Plan:</b> The default federal plan featuring fixed equal monthly
    payments for 120 months.<br/>
    &bull; <b>*Repayment Assistance Plan (RAP):</b> Available for Direct Student Loan borrowers
    under federal guidelines. RAP monthly payments are calculated based on prior-year Adjusted
    Gross Income (AGI) on a sliding scale (1% to 10% of AGI) and claimed dependents ($50/mo
    reduction per dependent), with a mandatory minimum payment of
    ${program_meta['rap_minimum_payment']:.2f}/month. RAP waives unpaid monthly interest for
    on-time payments, preventing runaway debt growth.{plus_extra_disc}<br/>
    &bull; <b>Official Simulators:</b> Visit <b>StudentAid.gov/loan-simulator</b> to compare
    official plans and apply with your federal loan servicer.
    </font>
    """
    disc_box = Table(
        [[Paragraph(repay_disclosures, ParagraphStyle(
            "DBox", fontName="Helvetica", fontSize=8, leading=10.5, textColor=DARK_GRAY,
        ))]],
        colWidths=[540],
    )
    disc_box.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT_BG),
        ("BOX", (0, 0), (-1, -1), 1, NAVY),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(disc_box)

doc.build(story)

# ---------------------------------------------------------------------------
# PACKET ASSEMBLY: MERGE COVER + SELECTED FLYERS
# ---------------------------------------------------------------------------

name_parts = student_name.split()
if len(name_parts) >= 2:
    last_name = name_parts[-1]
    first_name = " ".join(name_parts[:-1])
    formatted_name = f"{last_name}, {first_name}"
else:
    formatted_name = student_name

student_folder = f"{formatted_name} FA Info"
final_output_dir = os.path.join(OUTPUT_ROOT, student_folder)
os.makedirs(final_output_dir, exist_ok=True)

merger = PdfMerger()
merger.append(cover_pdf)
for flyer_name in selected_flyers:
    fp = FLYER_MAP.get(flyer_name)
    if fp and os.path.exists(fp):
        merger.append(fp)

output_filename = f"{formatted_name} FA Info.pdf"
final_pdf_path = os.path.join(final_output_dir, output_filename)

with open(final_pdf_path, "wb") as f_out:
    merger.write(f_out)
merger.close()

if os.path.exists(cover_pdf):
    os.remove(cover_pdf)

print(f"Success! Clock-hour packet generated at: {final_pdf_path}")